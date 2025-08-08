#!/usr/bin/env python3
"""
Marks Upload Template Generator
Creates Excel templates for bulk marks upload
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile


def create_marks_upload_template(students, subjects, total_marks=100, grade="", stream="", term="", assessment_type="", subject_total_marks=None):
    """
    Create a customized marks upload template based on students and subjects.
    
    Args:
        students: List of student objects
        subjects: List of subject names
        total_marks: Default total marks (default: 100)
        grade: Grade level
        stream: Stream name
        term: Term name
        assessment_type: Assessment type
        subject_total_marks: Dict of subject names to total marks
    
    Returns:
        str: Path to the created template file
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Upload Template"
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    info_font = Font(bold=True, color="2F75B5")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add template information
    ws['A1'] = f"Marks Upload Template - {grade} {stream}"
    ws['A1'].font = Font(bold=True, size=14, color="2F75B5")
    
    ws['A2'] = f"Term: {term}"
    ws['A2'].font = info_font
    
    ws['A3'] = f"Assessment Type: {assessment_type}"
    ws['A3'].font = info_font
    
    ws['A4'] = "Instructions: Enter marks for each student in the respective subject columns"
    ws['A4'].font = Font(italic=True, color="666666")
    
    # Start the data table at row 6
    start_row = 6
    
    # Create headers
    headers = ["Student Name", "Admission Number"] + subjects
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add max marks row
    max_marks_row = start_row + 1
    ws.cell(row=max_marks_row, column=1, value="Max Marks →").font = Font(bold=True, color="D9534F")
    ws.cell(row=max_marks_row, column=2, value="").font = Font(bold=True)
    
    for col_num, subject in enumerate(subjects, 3):
        max_mark = subject_total_marks.get(subject, total_marks) if subject_total_marks else total_marks
        cell = ws.cell(row=max_marks_row, column=col_num, value=max_mark)
        cell.font = Font(bold=True, color="D9534F")
        cell.fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
        cell.border = border
    
    # Add student data
    data_start_row = start_row + 2
    for row_num, student in enumerate(students, data_start_row):
        # Student name
        ws.cell(row=row_num, column=1, value=student.name).border = border
        
        # Admission number (if available)
        admission_num = getattr(student, 'admission_number', '') or getattr(student, 'student_id', '') or f"STU{student.id:04d}"
        ws.cell(row=row_num, column=2, value=admission_num).border = border
        
        # Subject columns (empty for user to fill)
        for col_num in range(3, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num, value="")
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add notes sheet
    notes_ws = wb.create_sheet("Instructions & Notes")
    
    # Instructions
    instructions = [
        ["📋 MARKS UPLOAD TEMPLATE INSTRUCTIONS"],
        [""],
        ["🎯 Purpose:", "Use this template to upload marks for multiple students at once"],
        [""],
        ["📝 How to Use:"],
        ["1.", "Enter marks for each student in the subject columns"],
        ["2.", "Do NOT modify the student names or admission numbers"],
        ["3.", "Do NOT modify the header row"],
        ["4.", "Leave cells empty if a student was absent (will be marked as 0)"],
        ["5.", "Ensure marks do not exceed the maximum shown in row 7"],
        [""],
        ["⚠️  Important Notes:"],
        ["•", "Only numeric values are allowed in mark columns"],
        ["•", "Marks must be between 0 and the maximum for each subject"],
        ["•", "Save the file as .xlsx format before uploading"],
        ["•", "Do not add or remove columns"],
        [""],
        ["🔍 Validation:"],
        ["•", "The system will validate all marks before saving"],
        ["•", "Any errors will be reported for correction"],
        ["•", "Successfully uploaded marks will be confirmed"],
        [""],
        [f"📊 Template Details:"],
        [f"Grade:", grade],
        [f"Stream:", stream],
        [f"Term:", term],
        [f"Assessment:", assessment_type],
        [f"Students:", len(students)],
        [f"Subjects:", len(subjects)]
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        for col_num, text in enumerate(instruction, 1):
            cell = notes_ws.cell(row=row_num, column=col_num, value=text)
            if row_num == 1:  # Title
                cell.font = Font(bold=True, size=16, color="2F75B5")
            elif col_num == 1 and text in ["🎯 Purpose:", "📝 How to Use:", "⚠️  Important Notes:", "🔍 Validation:", "📊 Template Details:"]:
                cell.font = Font(bold=True, color="2F75B5")
            elif text.startswith(("Grade:", "Stream:", "Term:", "Assessment:", "Students:", "Subjects:")):
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths for notes
    for column in notes_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        notes_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_dir = tempfile.gettempdir()
    filename = f"marks_template_{grade}_{stream}_{term}_{assessment_type}.xlsx"
    template_path = os.path.join(temp_dir, filename)
    
    wb.save(template_path)
    wb.close()
    
    return template_path


def create_empty_marks_template(education_level="upper_primary"):
    """
    Create a generic/empty marks upload template.
    
    Args:
        education_level: Education level for the template
    
    Returns:
        str: Path to the created template file
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Upload Template"
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    info_font = Font(bold=True, color="2F75B5")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add template information
    ws['A1'] = f"Generic Marks Upload Template - {education_level.replace('_', ' ').title()}"
    ws['A1'].font = Font(bold=True, size=14, color="2F75B5")
    
    ws['A2'] = "Instructions: Replace this with your actual class data"
    ws['A2'].font = info_font
    
    # Sample headers
    sample_subjects = ["Mathematics", "English", "Science", "Social Studies", "Kiswahili"]
    headers = ["Student Name", "Admission Number"] + sample_subjects
    
    # Create headers at row 4
    start_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add max marks row
    max_marks_row = start_row + 1
    ws.cell(row=max_marks_row, column=1, value="Max Marks →").font = Font(bold=True, color="D9534F")
    ws.cell(row=max_marks_row, column=2, value="").font = Font(bold=True)
    
    for col_num in range(3, len(headers) + 1):
        cell = ws.cell(row=max_marks_row, column=col_num, value=100)
        cell.font = Font(bold=True, color="D9534F")
        cell.fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
        cell.border = border
    
    # Add sample student data
    sample_students = [
        ("Sample Student 1", "STU001"),
        ("Sample Student 2", "STU002"),
        ("Sample Student 3", "STU003")
    ]
    
    data_start_row = start_row + 2
    for row_num, (name, admission) in enumerate(sample_students, data_start_row):
        ws.cell(row=row_num, column=1, value=name).border = border
        ws.cell(row=row_num, column=2, value=admission).border = border
        
        # Empty subject columns
        for col_num in range(3, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num, value="")
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to static templates directory
    static_dir = os.path.join(os.path.dirname(__file__))
    filename = f"marks_upload_template_example_{education_level}.xlsx"
    template_path = os.path.join(static_dir, filename)
    
    wb.save(template_path)
    wb.close()
    
    return template_path


if __name__ == "__main__":
    # Test the function
    create_empty_marks_template("upper_primary")
    print("Template created successfully!")
