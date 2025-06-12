#!/usr/bin/env python3
"""
Student upload template generator for the Hillview School Management System.
Creates Excel and CSV templates with enhanced format including grade and stream columns.
"""

import pandas as pd
import os
from datetime import datetime

def create_student_upload_template():
    """Create both Excel and CSV templates for student upload."""
    print("🔧 Creating enhanced student upload templates")
    print("=" * 50)
    
    # Get the directory where this script is located
    template_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Enhanced sample data with grade and stream columns
    sample_data = [
        {
            "name": "John Doe",
            "admission_number": "ADM001", 
            "grade": "Grade 4",
            "stream": "A",
            "gender": "Male"
        },
        {
            "name": "Jane Smith",
            "admission_number": "ADM002",
            "grade": "Grade 4", 
            "stream": "A",
            "gender": "Female"
        },
        {
            "name": "Michael Johnson",
            "admission_number": "ADM003",
            "grade": "Grade 4",
            "stream": "B", 
            "gender": "Male"
        },
        {
            "name": "Emily Brown",
            "admission_number": "ADM004",
            "grade": "Grade 5",
            "stream": "A",
            "gender": "Female"
        },
        {
            "name": "David Wilson",
            "admission_number": "ADM005",
            "grade": "Grade 5",
            "stream": "B",
            "gender": "Male"
        },
        {
            "name": "Sarah Davis",
            "admission_number": "ADM006",
            "grade": "Grade 6",
            "stream": "A", 
            "gender": "Female"
        },
        {
            "name": "James Miller",
            "admission_number": "",  # Example of missing admission number
            "grade": "Grade 6",
            "stream": "B",
            "gender": "Male"
        },
        {
            "name": "Lisa Anderson",
            "admission_number": "ADM008",
            "grade": "Grade 4",
            "stream": "A",
            "gender": "Female"
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(sample_data)
    
    # Create Excel template
    excel_path = os.path.join(template_dir, 'student_upload_template.xlsx')
    try:
        # Create Excel file with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write the main data
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Add formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F7D53", end_color="1F7D53", fill_type="solid")
            
            # Format header row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for col_num, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = max(len(str(column)), max(len(str(val)) for val in df[column]))
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 20)
            
            # Add instructions sheet
            instructions_data = [
                ["📋 Student Upload Template Instructions", ""],
                ["", ""],
                ["Required Columns:", ""],
                ["• name", "Student's full name (REQUIRED)"],
                ["", ""],
                ["Optional Columns:", ""],
                ["• admission_number", "Student's admission/registration number"],
                ["• grade", "Student's grade (e.g., 'Grade 4', 'Grade 5')"],
                ["• stream", "Student's stream (e.g., 'A', 'B', 'C')"],
                ["• gender", "Student's gender (Male/Female)"],
                ["", ""],
                ["Column Name Variations Supported:", ""],
                ["Name variations:", "name, student_name, full_name"],
                ["Admission variations:", "admission_number, addmission_number, adm_number, reg_number"],
                ["", ""],
                ["Important Notes:", ""],
                ["• Only 'name' column is required", ""],
                ["• Admission numbers will be auto-generated if missing", ""],
                ["• Grade and stream will auto-assign students to correct classes", ""],
                ["• System handles column name variations automatically", ""],
                ["• Remove this instructions sheet before uploading", ""],
                ["", ""],
                ["Example Data:", ""],
                ["See the 'Students' sheet for sample data format", ""]
            ]
            
            instructions_df = pd.DataFrame(instructions_data, columns=['Instruction', 'Details'])
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Format instructions sheet
            instructions_ws = writer.sheets['Instructions']
            
            # Title formatting
            title_font = Font(bold=True, size=14, color="1F7D53")
            instructions_ws.cell(row=1, column=1).font = title_font
            
            # Auto-adjust column widths for instructions
            for col_num in range(1, 3):
                column_letter = get_column_letter(col_num)
                instructions_ws.column_dimensions[column_letter].width = 40
        
        print(f"✅ Created Excel template: {excel_path}")
        
    except Exception as e:
        print(f"❌ Error creating Excel template: {e}")
        # Fallback to simple Excel creation
        df.to_excel(excel_path, index=False)
        print(f"✅ Created basic Excel template: {excel_path}")
    
    # Create CSV template
    csv_path = os.path.join(template_dir, 'student_upload_template.csv')
    df.to_csv(csv_path, index=False)
    print(f"✅ Created CSV template: {csv_path}")
    
    # Create a simple CSV with just headers for minimal template
    headers_only_data = [
        {
            "name": "Student Name Here",
            "admission_number": "ADM001 (optional)",
            "grade": "Grade 4 (optional)",
            "stream": "A (optional)",
            "gender": "Male/Female (optional)"
        }
    ]
    
    headers_df = pd.DataFrame(headers_only_data)
    headers_csv_path = os.path.join(template_dir, 'student_upload_template_headers_only.csv')
    headers_df.to_csv(headers_csv_path, index=False)
    print(f"✅ Created headers-only CSV: {headers_csv_path}")
    
    print(f"\n📊 Template Summary:")
    print(f"  - Columns: {list(df.columns)}")
    print(f"  - Sample records: {len(sample_data)}")
    print(f"  - Excel features: Formatting, instructions sheet")
    print(f"  - CSV features: Simple format, headers-only version")
    
    return excel_path, csv_path

def create_minimal_student_template():
    """Create a minimal template with just the required columns."""
    template_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Minimal data with just name and optional fields
    minimal_data = [
        {"name": "John Doe", "admission_number": "ADM001", "grade": "Grade 4", "stream": "A"},
        {"name": "Jane Smith", "admission_number": "ADM002", "grade": "Grade 4", "stream": "B"},
        {"name": "Michael Johnson", "admission_number": "", "grade": "Grade 5", "stream": "A"}
    ]
    
    df = pd.DataFrame(minimal_data)
    
    # Save minimal templates
    excel_path = os.path.join(template_dir, 'student_upload_template_minimal.xlsx')
    csv_path = os.path.join(template_dir, 'student_upload_template_minimal.csv')
    
    df.to_excel(excel_path, index=False)
    df.to_csv(csv_path, index=False)
    
    return excel_path, csv_path

if __name__ == '__main__':
    try:
        excel_file, csv_file = create_student_upload_template()
        minimal_excel, minimal_csv = create_minimal_student_template()
        
        print(f"\n🎉 Student upload templates created successfully!")
        print(f"📁 Full Excel template: {excel_file}")
        print(f"📁 Full CSV template: {csv_file}")
        print(f"📁 Minimal Excel template: {minimal_excel}")
        print(f"📁 Minimal CSV template: {minimal_csv}")
        
    except Exception as e:
        print(f"❌ Error creating templates: {e}")
        import traceback
        traceback.print_exc()
