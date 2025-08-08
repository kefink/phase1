"""
Marks Upload Template Generator
Creates Excel templates for bulk marks upload
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_marks_template(students, subjects, grade=None, stream=None, term=None, assessment_type=None, total_marks=100):
    """
    Create an Excel template for marks upload
    
    Args:
        students: List of student objects or dictionaries
        subjects: List of subject objects or dictionaries  
        grade: Grade level
        stream: Class stream
        term: Academic term
        assessment_type: Type of assessment
        total_marks: Maximum marks (default 100)
    
    Returns:
        BytesIO object containing the Excel file
    """
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Upload Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="268BD2", end_color="268BD2", fill_type="solid")
    info_font = Font(bold=True, color="2AA198")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add title and information
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "HILLVIEW SCHOOL - MARKS UPLOAD TEMPLATE"
    ws[f'A{current_row}'].font = Font(bold=True, size=16, color="268BD2")
    ws[f'A{current_row}'].alignment = center_alignment
    current_row += 2
    
    # Template information
    info_data = [
        ("Grade:", grade or "Not Specified"),
        ("Stream:", stream or "Not Specified"), 
        ("Term:", term or "Not Specified"),
        ("Assessment:", assessment_type or "Not Specified"),
        ("Max Marks:", total_marks),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M"))
    ]
    
    for label, value in info_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = str(value)
        ws[f'B{current_row}'].font = Font(bold=True)
        current_row += 1
    
    current_row += 1
    
    # Instructions
    instructions = [
        "INSTRUCTIONS:",
        "1. Enter marks for each student in the respective subject columns",
        "2. Marks should be numeric values only (0 to {})".format(total_marks),
        "3. Leave cells empty for absent students",
        "4. Do not modify the header row or student names",
        "5. Save file and upload using the bulk upload feature"
    ]
    
    for instruction in instructions:
        ws[f'A{current_row}'] = instruction
        if current_row == len(instructions) + current_row - len(instructions):  # First instruction
            ws[f'A{current_row}'].font = Font(bold=True, color="DC322F")
        else:
            ws[f'A{current_row}'].font = Font(color="586E75")
        current_row += 1
    
    current_row += 2
    
    # Create headers
    headers = ["Student Name", "Admission Number"]
    
    # Add subject headers
    for subject in subjects:
        if hasattr(subject, 'name'):
            subject_name = subject.name
        elif isinstance(subject, dict):
            subject_name = subject.get('name', str(subject))
        else:
            subject_name = str(subject)
        headers.append(f"{subject_name} (Max: {total_marks})")
    
    # Write headers
    header_row = current_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    # Add student rows
    for student in students:
        if hasattr(student, 'name'):
            student_name = student.name
            admission_number = getattr(student, 'admission_number', '') or getattr(student, 'id', '')
        elif isinstance(student, dict):
            student_name = student.get('name', str(student))
            admission_number = student.get('admission_number', '') or student.get('id', '')
        else:
            student_name = str(student)
            admission_number = ''
        
        # Student name
        ws.cell(row=current_row, column=1, value=student_name).border = border
        # Admission number  
        ws.cell(row=current_row, column=2, value=str(admission_number)).border = border
        
        # Add empty cells for marks (with borders)
        for col_num in range(3, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num, value="")
            cell.border = border
            cell.alignment = center_alignment
        
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for row in range(1, current_row):
            try:
                cell_value = str(ws[f'{column_letter}{row}'].value or '')
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        # Set minimum and maximum column widths
        adjusted_width = min(max(max_length + 2, 12), 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add footer note
    current_row += 2
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "Note: This template is automatically generated. Please do not modify the structure."
    ws[f'A{current_row}'].font = Font(italic=True, color="93A1A1")
    ws[f'A{current_row}'].alignment = center_alignment
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_simple_marks_template(grade=None, stream=None, term=None, assessment_type=None):
    """
    Create a simple template when student/subject data is not available
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Template"
    
    # Add basic structure
    ws['A1'] = "MARKS UPLOAD TEMPLATE"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = f"Grade: {grade or 'Not Specified'}"
    ws['A4'] = f"Stream: {stream or 'Not Specified'}"
    ws['A5'] = f"Term: {term or 'Not Specified'}"
    ws['A6'] = f"Assessment: {assessment_type or 'Not Specified'}"
    
    # Basic headers
    ws['A8'] = "Student Name"
    ws['B8'] = "Admission Number"
    ws['C8'] = "Subject 1 (Max: 100)"
    ws['D8'] = "Subject 2 (Max: 100)"
    ws['E8'] = "Subject 3 (Max: 100)"
    
    # Style headers
    for col in ['A8', 'B8', 'C8', 'D8', 'E8']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color="268BD2", end_color="268BD2", fill_type="solid")
    
    # Add sample rows
    sample_students = ["Student Name 1", "Student Name 2", "Student Name 3"]
    for i, student in enumerate(sample_students, 9):
        ws[f'A{i}'] = student
        ws[f'B{i}'] = f"ADM00{i-8}"
    
    # Auto-adjust columns
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
